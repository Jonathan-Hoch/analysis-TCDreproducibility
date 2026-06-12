"""
Build minute-level revised manuscript figures/tables.

Outputs are written to B_Revised_Figures only. This script uses:
- Baseline absolute values for baseline agreement/BA.
- Delta CPT 1min and Delta CPT 2min columns for minute-level CPT agreement/BA.
- Baseline delta = 0 for Fig 1 response time-course.
"""

from pathlib import Path
import os
import warnings

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import pandas as pd
import pingouin as pg
from scipy import stats
import statsmodels.formula.api as smf
from statsmodels.stats.anova import AnovaRM

warnings.filterwarnings("ignore")
np.random.seed(42)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJ_ROOT = Path(os.environ.get("TCD_PROJECT_ROOT", SCRIPT_DIR.parent))
ACQ_DIR = Path(os.environ.get("TCD_DATA_DIR", PROJ_ROOT / "ACQ files"))
FIG_DIR = Path(os.environ.get("TCD_OUTPUT_DIR", SCRIPT_DIR))
MASTER_XLSX = Path(os.environ.get("TCD_MASTER_XLSX", ACQ_DIR / "CPT Data_visit split.xlsx"))
ETCO2_XLSX = Path(os.environ.get("TCD_ETCO2_XLSX", ACQ_DIR / "CPT_ETCO2_Resp_Comparison.xlsx"))
FIG_DIR.mkdir(exist_ok=True)

FSU_GARNET = "#782F40"
FSU_GOLD = "#CEB888"
FSU_BLUE = "#425563"
VAULT_GARNET = "#A6192E"
TEAL = "#5CB8B2"
GRAY = "#666666"


def save_figure(fig, out, **kwargs):
    fig.savefig(out, **kwargs)
    if out.suffix.lower() == ".png":
        fig.savefig(out.with_suffix(".pdf"), **{k: v for k, v in kwargs.items() if k != "dpi"})


def save_png_as_pdf(png_path):
    png_path = Path(png_path)
    pdf_path = png_path.with_suffix(".pdf")
    if not png_path.exists():
        return
    img = plt.imread(png_path)
    height, width = img.shape[:2]
    fig_w = width / 500
    fig_h = height / 500
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.imshow(img)
    ax.axis("off")
    fig.savefig(pdf_path, bbox_inches="tight", pad_inches=0)
    plt.close(fig)


GLOBAL_EXCL_IDS = {45}
VISIT_EXCL = {(13, 1)}  # K32V1 ET-CO2 artifact, visit-level regression exclusion.
VAR_EXCL = {
    "etco2": {"F02V1", "M24V1", "K32V1/F10V1"},
    "mcav_peak": {"K10V1", "F02V1", "M24V1"},
    "mcav_min": {"K10V1", "F02V1", "M24V1"},
    "mcav_mean": {"K10V1", "F02V1", "M24V1"},
    "mcav_pulse": {"K10V1", "F02V1", "M24V1"},
    "mcav_gpi": {"K10V1", "F02V1", "M24V1"},
    "cvci": {"K10V1", "F02V1", "M24V1"},
    "cvri": {"M26V1", "F17V1", "F15V1", "M02V1", "F27V1/M16", "K05V1", "F26V2", "M21V1"},
    "smo2": {"K05V1"},
    "q": set(),
    "tpr": {"F05V1"},
    "sbp": set(),
    "dbp": set(),
    "mbp": set(),
    "hr": set(),
}

EPOCHS = ["Base", "1min", "2min"]
EPOCH_LABEL = {"Base": "Baseline", "1min": "Min 1", "2min": "Min 2"}
FIG_XLABELS = ["Baseline", "Min 1", "Min 2"]

ALL15 = [
    ("sbp", "SBP", "SBP", "mmHg"),
    ("dbp", "DBP", "DBP", "mmHg"),
    ("mbp", "MBP", "MBP", "mmHg"),
    ("hr", "HR", "HR", "bpm"),
    ("q", "Q", "Q (cardiac)", "L/min"),
    ("tpr", "TPR", "TPR", "dynes*s/cm5"),
    ("etco2", "ET-CO2", "ET-CO2", "mmHg"),
    ("mcav_peak", "MCAv peak", "MCAv peak", "cm/s"),
    ("mcav_min", "MCAv minimum", "MCAv min", "cm/s"),
    ("mcav_mean", "MCAv mean", "MCAv mean", "cm/s"),
    ("mcav_pulse", "MCAv pulse", "MCAv pulse", "cm/s"),
    ("mcav_gpi", "MCAv GPI", "MCAv pulsatility", "ratio"),
    ("cvci", "MCAv CVCi", "CVCi", "cm/s/mmHg"),
    ("cvri", "MCAv Resis", "CVRi", "mmHg*s/cm"),
    ("smo2", "SmO2", "SmO2", "%"),
]

FIG1_VARS = [
    ("sbp", "SBP", "Delta SBP (mmHg)"),
    ("dbp", "DBP", "Delta DBP (mmHg)"),
    ("mbp", "MBP", "Delta MBP (mmHg)"),
    ("hr", "HR", "Delta HR (bpm)"),
    ("mcav_mean", "MCAv mean", "Delta MCAv mean (cm/s)"),
    ("cvci", "MCAv CVCi", "Delta CVCi (cm/s/mmHg)"),
    ("mcav_gpi", "MCAv GPI", "Delta MCAv pulsatility"),
    ("etco2", "ET-CO2", "Delta ET-CO2 (mmHg)"),
]

FIG2_VARS = [
    ("sbp", "SBP", "SBP", "mmHg"),
    ("dbp", "DBP", "DBP", "mmHg"),
    ("mbp", "MBP", "MBP", "mmHg"),
    ("hr", "HR", "HR", "bpm"),
]

FIG3_VARS = [
    ("mcav_mean", "MCAv mean", "MCAv mean", "cm/s"),
    ("cvci", "MCAv CVCi", "CVCi", "cm/s/mmHg"),
    ("mcav_gpi", "MCAv GPI", "MCAv pulsatility", "ratio"),
    ("etco2", "ET-CO2", "ET-CO2", "mmHg"),
]


def load_data():
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in ws[1]]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    df = pd.DataFrame(rows, columns=headers)
    df.replace(["NaN", "nan"], np.nan, inplace=True)

    wb_p = openpyxl.load_workbook(ETCO2_XLSX, data_only=True)
    ws_p = wb_p["ETCO2_Resp_Comparison"]
    ph = [c.value for c in ws_p[1]]
    pr = [list(r) for r in ws_p.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    patch_df = pd.DataFrame(pr, columns=ph)
    patch_df.replace(["NaN", "nan"], np.nan, inplace=True)
    patch_ids = {"K22V1", "K27V1", "K32V1/F10V1", "K36V1"}
    patched_n = 0
    for _, patch_row in patch_df.iterrows():
        sid = patch_row.get("Subject ID")
        if sid not in patch_ids:
            continue
        mask = df["1 Subject ID"] == sid
        for col in [c for c in patch_df.columns if c not in ("Subject ID", "Source")]:
            if col in df.columns and mask.any():
                cur = df.loc[mask, col].iloc[0]
                val = patch_row[col]
                if pd.isna(cur) and pd.notna(val):
                    df.loc[mask, col] = val
                    patched_n += 1
    print(f"Loaded {df.shape[0]} rows x {df.shape[1]} columns. Patched {patched_n} ET-CO2 cells.")
    return df


def to_num(x):
    return pd.to_numeric(x, errors="coerce")


def value_from_row(row, visit, stem, epoch, delta_baseline=False):
    if epoch == "Base":
        if delta_baseline:
            return 0.0
        return to_num(row.get(f"{visit} {stem} Base"))
    return to_num(row.get(f"{visit} Delta {stem} CPT {epoch}"))


def include_subject(row, var_key):
    pid = int(row["1 Identifier"])
    sid = str(row.get("1 Subject ID", ""))
    return pid not in GLOBAL_EXCL_IDS and sid not in VAR_EXCL.get(var_key, set())


def build_paired(df, var_key, stem, epoch, delta_baseline=False, subset=None):
    rows = []
    for _, row in df.iterrows():
        if not include_subject(row, var_key):
            continue
        if subset is not None and not subset(row):
            continue
        pid = int(row["1 Identifier"])
        sid = str(row.get("1 Subject ID", ""))
        v1 = value_from_row(row, 1, stem, epoch, delta_baseline=delta_baseline)
        v2 = value_from_row(row, 2, stem, epoch, delta_baseline=delta_baseline)
        if pd.notna(v1) and pd.notna(v2):
            rows.append({"pid": pid, "sid": sid, "v1": float(v1), "v2": float(v2)})
    return pd.DataFrame(rows)


def long_for_var(df, var_key, stem, delta_baseline=True, subset=None):
    rows = []
    for _, row in df.iterrows():
        if not include_subject(row, var_key):
            continue
        if subset is not None and not subset(row):
            continue
        pid = int(row["1 Identifier"])
        for visit in (1, 2):
            for epoch in EPOCHS:
                val = value_from_row(row, visit, stem, epoch, delta_baseline=delta_baseline)
                if pd.notna(val):
                    rows.append({"pid": pid, "visit": f"V{visit}", "time": EPOCH_LABEL[epoch], "epoch": epoch, "value": float(val)})
    return pd.DataFrame(rows)


def lin_ccc(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    mx, my = np.mean(x), np.mean(y)
    vx, vy = np.var(x, ddof=0), np.var(y, ddof=0)
    cov_xy = np.cov(x, y, ddof=0)[0, 1]
    denom = vx + vy + (mx - my) ** 2
    return np.nan if denom == 0 else (2 * cov_xy) / denom


def icc_ccc(dp):
    if len(dp) < 4:
        return {"ICC3k": np.nan, "ICC3k_lo": np.nan, "ICC3k_hi": np.nan, "ICC3k_p": np.nan, "CCC": np.nan}
    long = pd.concat(
        [
            dp[["pid", "v1"]].rename(columns={"v1": "val"}).assign(rater=1),
            dp[["pid", "v2"]].rename(columns={"v2": "val"}).assign(rater=2),
        ],
        ignore_index=True,
    )
    try:
        icc_res = pg.intraclass_corr(data=long, targets="pid", raters="rater", ratings="val")
        row = icc_res[icc_res["Type"] == "ICC(C,k)"].iloc[0]
        ci = row["CI95"]
        return {
            "ICC3k": float(row["ICC"]),
            "ICC3k_lo": float(ci[0]),
            "ICC3k_hi": float(ci[1]),
            "ICC3k_p": float(row["pval"]),
            "CCC": float(lin_ccc(dp["v1"], dp["v2"])),
        }
    except Exception:
        return {"ICC3k": np.nan, "ICC3k_lo": np.nan, "ICC3k_hi": np.nan, "ICC3k_p": np.nan, "CCC": float(lin_ccc(dp["v1"], dp["v2"]))}


def ba_stats(dp):
    n = len(dp)
    if n < 4:
        return {"n": n}
    mean_val = (dp["v1"] + dp["v2"]) / 2
    diff_val = dp["v1"] - dp["v2"]
    md = float(diff_val.mean())
    sd = float(diff_val.std(ddof=1))
    sl, ic, r_pb, p_pb, _ = stats.linregress(mean_val, diff_val)
    use_pb = p_pb < 0.05
    fitted = (ic + sl * mean_val.values) if use_pb else np.full(n, md)
    resid = diff_val.values - fitted
    r_hs, p_hs = stats.spearmanr(mean_val, np.abs(resid))
    use_hs = p_hs < 0.05
    x_mid = float(mean_val.mean())
    mid = float(ic + sl * x_mid) if use_pb else md
    if use_hs:
        sl_sd, ic_sd, _, _, _ = stats.linregress(mean_val.values, np.abs(resid))
        floor = max(float(np.abs(resid).mean()) * 0.25, 1e-6)
        sd_mid = max(ic_sd + sl_sd * x_mid, floor)
        loa_lo = mid - 1.96 * sd_mid
        loa_hi = mid + 1.96 * sd_mid
        loa_type = "regression+fan" if use_pb else "fan"
    elif use_pb:
        sdr = float(np.std(resid, ddof=2))
        loa_lo = mid - 1.96 * sdr
        loa_hi = mid + 1.96 * sdr
        loa_type = "regression"
    else:
        loa_lo = md - 1.96 * sd
        loa_hi = md + 1.96 * sd
        loa_type = "standard"
    return {
        "n": n,
        "Mean_diff": md,
        "SD_diff": sd,
        "LoA_type": loa_type,
        "LoA_lo": loa_lo,
        "LoA_hi": loa_hi,
        "Prop_bias_r": float(r_pb),
        "Prop_bias_P": float(p_pb),
        "Hetero_rho": float(r_hs),
        "Hetero_P": float(p_hs),
    }


def ba_line_values(dp, x):
    b = ba_stats(dp)
    mean_val = (dp["v1"] + dp["v2"]) / 2
    diff_val = dp["v1"] - dp["v2"]
    sl, ic, _r_pb, _p_pb, _ = stats.linregress(mean_val, diff_val)

    if b["LoA_type"] == "standard":
        bias = np.full_like(x, b["Mean_diff"], dtype=float)
        loa_lo = np.full_like(x, b["LoA_lo"], dtype=float)
        loa_hi = np.full_like(x, b["LoA_hi"], dtype=float)
    else:
        bias = ic + sl * x if "regression" in b["LoA_type"] else np.full_like(x, b["Mean_diff"], dtype=float)
        fitted = ic + sl * mean_val.values if "regression" in b["LoA_type"] else np.full(len(dp), b["Mean_diff"])
        resid = diff_val.values - fitted
        if "fan" in b["LoA_type"]:
            sl_sd, ic_sd, _, _, _ = stats.linregress(mean_val.values, np.abs(resid))
            floor = max(float(np.abs(resid).mean()) * 0.25, 1e-6)
            sd_line = np.maximum(ic_sd + sl_sd * x, floor)
        else:
            sd_line = np.full_like(x, float(np.std(resid, ddof=2)), dtype=float)
        loa_lo = bias - 1.96 * sd_line
        loa_hi = bias + 1.96 * sd_line
    return b, bias, loa_lo, loa_hi


def fmt_p(p):
    if pd.isna(p):
        return ""
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def summary_string(x):
    x = pd.Series(x).dropna().astype(float)
    if len(x) < 3:
        return ""
    p_norm = stats.shapiro(x).pvalue if len(x) <= 5000 else np.nan
    if pd.notna(p_norm) and p_norm >= 0.05:
        return f"{x.mean():.3g} +/- {x.std(ddof=1):.3g}", "mean_sd"
    return f"{x.median():.3g} [{(x.quantile(.75)-x.quantile(.25)):.3g}]", "median_iqr"


def paired_rank_biserial(diff):
    diff = pd.Series(diff).dropna().astype(float)
    diff = diff[diff != 0]
    if len(diff) == 0:
        return np.nan
    ranks = stats.rankdata(np.abs(diff), method="average")
    pos = float(ranks[diff.to_numpy() > 0].sum())
    neg = float(ranks[diff.to_numpy() < 0].sum())
    total = pos + neg
    return (pos - neg) / total if total else np.nan


def paired_comparison(dp):
    if len(dp) < 4:
        return {}
    diff = dp["v1"] - dp["v2"]
    try:
        norm_p = stats.shapiro(diff).pvalue
    except Exception:
        norm_p = np.nan
    if pd.notna(norm_p) and norm_p >= 0.05:
        t = stats.ttest_rel(dp["v1"], dp["v2"], nan_policy="omit")
        test = "paired t"
        p = float(t.pvalue)
        effect = float(diff.mean() / diff.std(ddof=1)) if diff.std(ddof=1) else np.nan
        effect_type = "paired d"
    else:
        try:
            w = stats.wilcoxon(dp["v1"], dp["v2"], zero_method="wilcox")
            p = float(w.pvalue)
        except Exception:
            p = np.nan
        test = "Wilcoxon"
        effect = float(paired_rank_biserial(diff))
        effect_type = "rank biserial"
    return {"Test": test, "P": p, "Effect": effect, "Effect_type": effect_type}


def agreement_row(df, var_key, stem, label, epoch):
    dp = build_paired(df, var_key, stem, epoch, delta_baseline=False)
    b = ba_stats(dp)
    icc = icc_ccc(dp)
    comp = paired_comparison(dp)
    if len(dp):
        v1s, v1_kind = summary_string(dp["v1"])
        v2s, v2_kind = summary_string(dp["v2"])
        diffs, diff_kind = summary_string(dp["v1"] - dp["v2"])
        mae = float(np.mean(np.abs(dp["v1"] - dp["v2"])))
        denom = np.abs((dp["v1"] + dp["v2"]) / 2)
        mape = float(np.mean(np.abs(dp["v1"] - dp["v2"]) / denom.replace(0, np.nan)) * 100)
        cv = float(np.std(dp["v1"] - dp["v2"], ddof=1) / np.mean((dp["v1"] + dp["v2"]) / 2) * 100)
    else:
        v1s = v2s = diffs = v1_kind = v2_kind = diff_kind = ""
        mae = mape = cv = np.nan
    row = {
        "Variable": label,
        "Epoch": EPOCH_LABEL[epoch],
        "n": len(dp),
        "Visit_1": v1s,
        "Visit_1_summary": v1_kind,
        "Visit_2": v2s,
        "Visit_2_summary": v2_kind,
        "Fixed_Bias": diffs,
        "Fixed_Bias_summary": diff_kind,
        "Comparison_Test": comp.get("Test", ""),
        "Comparison_P": comp.get("P", np.nan),
        "Effect": comp.get("Effect", np.nan),
        "Effect_type": comp.get("Effect_type", ""),
        "MAE": mae,
        "MAPE_%": mape,
        "CV_%": cv,
    }
    row.update(b)
    row.update(icc)
    return row


def draw_panel_label(ax, label):
    ax.text(-0.16, 1.13, label, transform=ax.transAxes, fontsize=11, fontweight="bold",
            va="bottom", ha="left", clip_on=False)


def padded_limits(values, pad_frac=0.08):
    vals = np.asarray(values, dtype=float)
    vals = vals[np.isfinite(vals)]
    if len(vals) == 0:
        return None
    lo = float(vals.min())
    hi = float(vals.max())
    if lo == hi:
        pad = max(abs(lo) * pad_frac, 1.0)
    else:
        pad = (hi - lo) * pad_frac
    return lo - pad, hi + pad


def ba_panel_values(dp):
    if len(dp) < 4:
        return None
    mean_val = (dp["v1"] + dp["v2"]) / 2
    diff_val = dp["v1"] - dp["v2"]
    x = np.linspace(float(mean_val.min()), float(mean_val.max()), 100)
    _b, bias, loa_lo, loa_hi = ba_line_values(dp, x)
    return mean_val, diff_val, x, bias, loa_lo, loa_hi


def ba_row_ylim(dps):
    y_vals = []
    for dp in dps:
        vals = ba_panel_values(dp)
        if vals is None:
            continue
        _mean_val, diff_val, _x, bias, loa_lo, loa_hi = vals
        y_vals.extend(diff_val.to_numpy(dtype=float))
        y_vals.extend(np.asarray(bias, dtype=float))
        y_vals.extend(np.asarray(loa_lo, dtype=float))
        y_vals.extend(np.asarray(loa_hi, dtype=float))
    return padded_limits(y_vals)


def add_ba_panel(ax, dp, title, ylabel=False, xlabel=False, xlim=None, ylim=None):
    if len(dp) < 4:
        ax.set_title(f"{title}\nn={len(dp)}")
        if xlim is not None:
            ax.set_xlim(*xlim)
        if ylim is not None:
            ax.set_ylim(*ylim)
        return
    mean_val, diff_val, x, bias, loa_lo, loa_hi = ba_panel_values(dp)
    ax.axhline(0, color="#999999", lw=0.8, ls="--")
    ax.plot(x, bias, color=FSU_GARNET, lw=1.3)
    ax.plot(x, loa_hi, color=FSU_GOLD, lw=1.0, ls="--")
    ax.plot(x, loa_lo, color=FSU_GOLD, lw=1.0, ls="--")
    ax.scatter(mean_val, diff_val, s=18, color=FSU_GARNET, alpha=0.7, edgecolors="none")
    ax.set_title(f"{title}\nn={len(dp)}", fontsize=8, pad=3)
    if ylabel:
        ax.set_ylabel("V1 - V2", fontsize=8)
    if xlabel:
        ax.set_xlabel("Mean (V1, V2)", fontsize=8)
    if xlim is not None:
        ax.set_xlim(*xlim)
    if ylim is not None:
        ax.set_ylim(*ylim)
    ax.tick_params(labelsize=7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)


def ba_grid(df, vars_, out_name, title):
    fig, axes = plt.subplots(len(vars_), len(EPOCHS), figsize=(9.0, 10.5))
    plt.subplots_adjust(hspace=0.66, wspace=0.36, left=0.15, right=0.98, top=0.92, bottom=0.07)
    labels = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    k = 0
    for r, (var_key, stem, label, units) in enumerate(vars_):
        row_dps = [build_paired(df, var_key, stem, epoch, delta_baseline=False) for epoch in EPOCHS]
        ylim = ba_row_ylim(row_dps)
        for c, epoch in enumerate(EPOCHS):
            ax = axes[r, c]
            dp = row_dps[c]
            add_ba_panel(ax, dp, EPOCH_LABEL[epoch], ylabel=(c == 0), xlabel=(r == len(vars_) - 1),
                         ylim=ylim)
            draw_panel_label(ax, labels[k])
            k += 1
            if c == 0:
                ax.text(-0.33, 0.5, f"{label}\n({units})", transform=ax.transAxes, rotation=90,
                        ha="center", va="center", fontsize=8, fontweight="bold")
    fig.suptitle(title, fontsize=12, fontweight="bold")
    out = FIG_DIR / out_name
    save_figure(fig, out, dpi=500, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out.name}")


def fig1_timecourse(df):
    rows = []
    fig, axes = plt.subplots(4, 2, figsize=(9, 13))
    plt.subplots_adjust(hspace=0.62, wspace=0.36, left=0.12, right=0.95, top=0.93, bottom=0.07)
    x = np.arange(len(EPOCHS))
    panel_labels = list("ABCDEFGH")
    for i, (var_key, stem, ylabel) in enumerate(FIG1_VARS):
        ax = axes.flat[i]
        long = long_for_var(df, var_key, stem, delta_baseline=True)
        for visit, color, marker, ls, dx in [("V1", FSU_GARNET, "o", "-", -0.07), ("V2", FSU_BLUE, "s", "--", 0.07)]:
            vals = []
            los = []
            his = []
            kinds = []
            for epoch in EPOCHS:
                y = long[(long["visit"] == visit) & (long["epoch"] == epoch)]["value"].astype(float)
                if len(y) == 0:
                    vals.append(np.nan); los.append(0); his.append(0); kinds.append("")
                    continue
                s, kind = summary_string(y)
                kinds.append(kind)
                if kind == "mean_sd":
                    center = y.mean(); err_lo = y.std(ddof=1); err_hi = err_lo
                else:
                    center = y.median(); err_lo = center - y.quantile(.25); err_hi = y.quantile(.75) - center
                vals.append(center); los.append(err_lo); his.append(err_hi)
                if epoch != "Base":
                    jitter = np.random.normal(dx, 0.025, len(y))
                    ax.scatter(np.full(len(y), EPOCHS.index(epoch)) + jitter, y, s=9, color=color, alpha=0.14, edgecolors="none")
            ax.errorbar(x + dx, vals, yerr=[los, his], fmt=marker + ls, color=color, lw=1.8, ms=5, capsize=3, label=visit)
            for epoch, center, lo, hi, kind in zip(EPOCHS, vals, los, his, kinds):
                rows.append({"Variable": ylabel.replace("Delta ", "").split(" (")[0], "Visit": visit, "Epoch": EPOCH_LABEL[epoch],
                             "Center": center, "Err_lo": lo, "Err_hi": hi, "Summary": kind})
        ax.axhline(0, color="#999999", lw=0.8, ls="--")
        ax.set_xticks(x)
        ax.set_xticklabels(FIG_XLABELS, fontsize=8)
        ax.set_ylabel(ylabel, fontsize=8)
        ax.set_title(ylabel.replace("Delta ", "").split(" (")[0], fontsize=9, fontweight="bold")
        ax.text(0.98, 0.98, f"n={long['pid'].nunique()}", transform=ax.transAxes, va="top", ha="right", fontsize=7, color=GRAY)
        draw_panel_label(ax, panel_labels[i])
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
    handles, labels = axes.flat[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper right", bbox_to_anchor=(0.99, 0.99), fontsize=8)
    fig.suptitle("CPT response: baseline reference with minute 1 and minute 2 deltas", fontsize=12, fontweight="bold")
    out = FIG_DIR / "Fig1_CPT_TimeCourse.png"
    save_figure(fig, out, dpi=500, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    pd.DataFrame(rows).to_csv(FIG_DIR / "Fig1_TimeCourse_Summary.csv", index=False)
    print(f"Saved: {out.name}")
    print("Saved: Fig1_TimeCourse_Summary.csv")


def fig1_anova(df):
    rows = []
    for var_key, stem, ylabel in FIG1_VARS:
        long = long_for_var(df, var_key, stem, delta_baseline=True)
        long = long.dropna(subset=["value"]).copy()
        var_label = ylabel.replace("Delta ", "").split(" (")[0]
        n_all = long["pid"].nunique()
        complete_pids = (
            long.drop_duplicates(["pid", "visit", "time"])
            .groupby("pid")
            .size()
            .loc[lambda s: s == 6]
            .index
        )
        long = long[long["pid"].isin(complete_pids)].copy()
        try:
            res = AnovaRM(long, depvar="value", subject="pid", within=["visit", "time"], aggregate_func="mean").fit()
            tab = res.anova_table.reset_index().rename(columns={"index": "Effect"})
            for _, r in tab.iterrows():
                f_val = r.get("F Value", np.nan)
                num_df = r.get("Num DF", np.nan)
                den_df = r.get("Den DF", np.nan)
                partial_eta_sq = (
                    (f_val * num_df) / ((f_val * num_df) + den_df)
                    if pd.notna(f_val) and pd.notna(num_df) and pd.notna(den_df)
                    else np.nan
                )
                rows.append({
                    "Variable": var_label,
                    "Test": "two-way repeated-measures ANOVA",
                    "Effect": r["Effect"],
                    "F": f_val,
                    "Num DF": num_df,
                    "Den DF": den_df,
                    "P": r.get("Pr > F", np.nan),
                    "partial_eta_sq": partial_eta_sq,
                    "n_participants": long["pid"].nunique(),
                    "n_participants_available": n_all,
                    "n_observations": len(long),
                })
        except Exception as exc:
            rows.append({"Variable": var_label, "Test": "ANOVA failed", "Effect": "", "P": np.nan,
                         "n_participants": long["pid"].nunique(), "n_participants_available": n_all,
                         "n_observations": len(long), "Error": str(exc)})
    out = pd.DataFrame(rows)
    out.to_csv(FIG_DIR / "Fig1_ANOVA_Time_Visit_Results.csv", index=False)
    sig = out[pd.to_numeric(out["P"], errors="coerce") < 0.05].copy()
    print("Saved: Fig1_ANOVA_Time_Visit_Results.csv")
    print("Significant Fig 1 ANOVA effects (P < 0.05):")
    if len(sig):
        print(sig[["Variable", "Effect", "F", "P", "n_participants"]].to_string(index=False))
    else:
        print("  None")


def build_tables_and_stats(df):
    rel_rows = []
    stats_rows = []
    for var_key, stem, label, _units in ALL15:
        for epoch in EPOCHS:
            row = agreement_row(df, var_key, stem, label, epoch)
            rel_rows.append(row)
            stats_rows.append({k: row.get(k, np.nan) for k in [
                "Variable", "Epoch", "n", "Mean_diff", "SD_diff", "LoA_type", "LoA_lo", "LoA_hi",
                "Prop_bias_r", "Prop_bias_P", "Hetero_rho", "Hetero_P",
                "ICC3k", "ICC3k_lo", "ICC3k_hi", "ICC3k_p", "CCC"
            ]})
    rel = pd.DataFrame(rel_rows)
    base = rel[rel["Epoch"] == "Baseline"].copy()
    mins = rel[rel["Epoch"].isin(["Min 1", "Min 2"])].copy()
    base.to_csv(FIG_DIR / "Table2_Baseline_All15_Reliability.csv", index=False)
    mins.to_csv(FIG_DIR / "Table3_Min1_Min2_All15_Reliability.csv", index=False)
    stats_df = pd.DataFrame(stats_rows)
    stats_df.to_csv(FIG_DIR / "BA_ICC_CCC_Statistics.csv", index=False)
    stats_df.to_csv(FIG_DIR / "SupplementaryTable_BA_Statistics.csv", index=False)
    print("Saved: Table2_Baseline_All15_Reliability.csv")
    print("Saved: Table3_Min1_Min2_All15_Reliability.csv")
    print("Saved: BA_ICC_CCC_Statistics.csv")
    print("Saved: SupplementaryTable_BA_Statistics.csv")
    missing = stats_df[stats_df[["ICC3k", "CCC"]].isna().any(axis=1)]
    if len(missing):
        print("Rows missing ICC/CCC:")
        print(missing[["Variable", "Epoch", "n", "ICC3k", "CCC"]].to_string(index=False))
    return stats_df


def build_reliability_stats_for_vars(df, vars_, subset=None):
    rows = []
    for var_key, stem, label, _units in vars_:
        for epoch in EPOCHS:
            dp = build_paired(df, var_key, stem, epoch, delta_baseline=False, subset=subset)
            vals = ba_stats(dp)
            vals.update(icc_ccc(dp))
            vals.update({"Variable": label, "Epoch": EPOCH_LABEL[epoch], "n": len(dp)})
            rows.append(vals)
    return pd.DataFrame(rows)


def exclude_unmatched_females(row):
    sex = str(row.get("1 Sex", ""))
    match = str(row.get("Menstrual phase matched?", ""))
    return not (sex == "Female" and "Un-matched" in match)


REL_STYLES = {
    "MCAv mean": (FSU_GARNET, "o", "-"),
    "CVCi": (FSU_GOLD, "o", "-"),
    "MCAv pulsatility": (TEAL, "o", "--"),
    "MBP": (FSU_BLUE, "s", "-"),
    "SBP": ("#7BAFD4", "s", "--"),
    "DBP": (VAULT_GARNET, "s", ":"),
    "HR": ("#572932", "D", "-"),
    "ET-CO2": ("#FFC72C", "D", "--"),
}


def add_reliability_bands(ax, stat, x_right, y_min=-0.10, show_labels=True):
    if stat == "ICC3k":
        bands = [
            (y_min, 0.50, "#FFCCCC", "Poor"),
            (0.50, 0.75, "#FFE0B2", "Moderate"),
            (0.75, 0.90, "#FFFDE7", "Good"),
            (0.90, 1.05, "#DCEDC8", "Excellent"),
        ]
        cuts = [0.50, 0.75, 0.90]
    else:
        bands = [
            (y_min, 0.20, "#FFCCCC", "Poor"),
            (0.20, 0.60, "#FFE0B2", "Fair"),
            (0.60, 0.70, "#FFFDE7", "Moderate"),
            (0.70, 0.90, "#DCEDC8", "Strong"),
            (0.90, 1.05, "#C8E6C9", "Very Strong"),
        ]
        cuts = [0.20, 0.60, 0.70, 0.90]
    for lo, hi, color, band_label in bands:
        ax.axhspan(lo, hi, alpha=0.20, color=color, zorder=0)
        if show_labels:
            ax.text(x_right, (lo + hi) / 2, band_label, fontsize=7, color="#444444",
                    ha="left", va="center", style="italic", clip_on=False)
    for cut in cuts:
        ax.axhline(cut, color="#888888", lw=0.7, ls=":", zorder=1)
    ax.axhline(0, color="#dddddd", lw=0.7, zorder=1)


def reliability_summary_plot(stats_df, vars_, out_name, title):
    plot_df = stats_df[stats_df["Variable"].isin([v[2] for v in vars_])].copy()
    var_order = [v[2] for v in vars_]
    fig, axes = plt.subplots(2, 1, figsize=(9.0, 10.5), sharex=True)
    plt.subplots_adjust(hspace=0.18, left=0.12, right=0.87, top=0.90, bottom=0.13)
    x = np.arange(len(EPOCHS))
    ep_map = {EPOCH_LABEL[e]: i for i, e in enumerate(EPOCHS)}
    for ax, stat, label in [(axes[0], "ICC3k", "ICC(3,k)"), (axes[1], "CCC", "CCC")]:
        add_reliability_bands(ax, stat, x_right=len(EPOCHS) - 0.34, y_min=-0.10, show_labels=True)
        for i, var in enumerate(var_order):
            d = plot_df[plot_df["Variable"] == var].copy()
            d["x"] = d["Epoch"].map(ep_map)
            d = d.sort_values("x")
            off = (i - len(var_order) / 2) * 0.055
            color, marker, ls = REL_STYLES[var]
            ax.plot(d["x"] + off, d[stat], marker=marker, ls=ls, lw=1.7, ms=5.5,
                    color=color, alpha=0.90, label=var, zorder=3)
            if stat == "ICC3k":
                ax.vlines(d["x"] + off, d["ICC3k_lo"], d["ICC3k_hi"],
                          color=color, lw=0.9, alpha=0.40, zorder=2)
        ax.set_ylabel(label, fontsize=10)
        ax.set_ylim(-0.1, 1.05)
        ax.set_xlim(-0.50, len(EPOCHS) - 0.20)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
    axes[0].set_title("A  ICC(3,k) [95% CI]\n(Koo & Li, 2016)", loc="left", fontweight="bold")
    axes[1].set_title("B  Lin's CCC\n(Akoglu, 2018)", loc="left", fontweight="bold")
    axes[1].set_xticks(x)
    axes[1].set_xticklabels(FIG_XLABELS, rotation=30, ha="right")
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", ncol=4, fontsize=8, framealpha=0.9,
               handlelength=2.0, markerscale=1.1)
    fig.suptitle(title, fontsize=12, fontweight="bold")
    out = FIG_DIR / out_name
    save_figure(fig, out, dpi=500, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out.name}")


def group_subset(group_name):
    def _subset(row):
        sex = str(row.get("1 Sex", ""))
        match = str(row.get("Menstrual phase matched?", ""))
        if group_name == "Male":
            return sex == "Male"
        if group_name == "Female matched":
            return sex == "Female" and "Un-matched" not in match
        if group_name == "Female unmatched":
            return sex == "Female" and "Un-matched" in match
        return False
    return _subset


def sex_reliability(df, vars_):
    rows = []
    groups = ["Male", "Female matched", "Female unmatched"]
    for group in groups:
        subset = group_subset(group)
        for var_key, stem, label, _units in vars_:
            for epoch in EPOCHS:
                dp = build_paired(df, var_key, stem, epoch, delta_baseline=False, subset=subset)
                vals = icc_ccc(dp)
                vals.update({"Group": group, "Variable": label, "Epoch": EPOCH_LABEL[epoch], "n": len(dp)})
                rows.append(vals)
    out = pd.DataFrame(rows)
    out.to_csv(FIG_DIR / "Fig5_ICC_CCC_Sex_Menstrual_Stats.csv", index=False)
    print("Saved: Fig5_ICC_CCC_Sex_Menstrual_Stats.csv")
    return out


def sex_reliability_plot(sex_df, vars_):
    groups = ["Male", "Female matched", "Female unmatched"]
    var_order = [v[2] for v in vars_]
    point_min = float(np.nanmin([sex_df["ICC3k"].min(), sex_df["CCC"].min(), -0.10]))
    y_min = max(-1.40, np.floor((point_min - 0.05) * 10) / 10)
    fig, axes = plt.subplots(2, 3, figsize=(14, 9.2), sharex=True, sharey=True)
    plt.subplots_adjust(hspace=0.18, wspace=0.12, left=0.08, right=0.91, top=0.86, bottom=0.23)
    x = np.arange(len(EPOCHS))
    ep_map = {EPOCH_LABEL[e]: i for i, e in enumerate(EPOCHS)}
    panel = list("ABCDEF")
    k = 0
    for c, group in enumerate(groups):
        for r, (stat, ylabel) in enumerate([("ICC3k", "ICC(3,k)"), ("CCC", "CCC")]):
            ax = axes[r, c]
            add_reliability_bands(ax, stat, x_right=len(EPOCHS) - 0.28,
                                  y_min=y_min, show_labels=(c == len(groups) - 1))
            gd = sex_df[sex_df["Group"] == group].copy()
            for i, var in enumerate(var_order):
                d = gd[gd["Variable"] == var].copy()
                d["x"] = d["Epoch"].map(ep_map)
                d = d.sort_values("x")
                off = (i - len(var_order) / 2) * 0.055
                color, marker, ls = REL_STYLES[var]
                ax.plot(d["x"] + off, d[stat], marker=marker, ls=ls, lw=1.25, ms=4.5,
                        color=color, alpha=0.90, label=var, zorder=3)
                if stat == "ICC3k":
                    ax.vlines(d["x"] + off, d["ICC3k_lo"], d["ICC3k_hi"],
                              color=color, lw=0.7, alpha=0.35, zorder=2)
            ax.set_title((group if r == 0 else "") + (f"\n{panel[k]}" if r == 0 else panel[k]),
                         fontsize=10, fontweight="bold")
            ax.set_ylim(y_min, 1.05)
            ax.set_xlim(-0.50, len(EPOCHS) - 0.12)
            if c == 0:
                ax.set_ylabel(ylabel)
            if r == 1:
                ax.set_xticks(x)
                ax.set_xticklabels(FIG_XLABELS, rotation=30, ha="right")
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            k += 1
    handles, labels = axes[0, 0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", bbox_to_anchor=(0.5, 0.02),
               ncol=4, fontsize=8, framealpha=0.9,
               handlelength=2.0, markerscale=1.1)
    fig.text(0.08, 0.115,
             "Negative reliability point estimates are shown; extremely wide ICC confidence intervals are clipped at the axis limit.",
             fontsize=7.5, color="#555555")
    fig.suptitle("Reliability by sex and menstrual-cycle matching", fontsize=12, fontweight="bold")
    out = FIG_DIR / "Fig5_ICC_CCC_Sex_Menstrual.png"
    save_figure(fig, out, dpi=500, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Saved: {out.name}")


def export_loo_influence_table(df, vars_):
    groups = ["All", "Female unmatched", "Exclude unmatched females"]

    def subset_for(group):
        if group == "All":
            return None
        if group == "Exclude unmatched females":
            return exclude_unmatched_females
        return group_subset(group)

    summary_rows = []
    loo_rows = []
    for group in groups:
        subset = subset_for(group)
        for var_key, stem, label, _units in vars_:
            for epoch in EPOCHS:
                dp = build_paired(df, var_key, stem, epoch, delta_baseline=False, subset=subset)
                full = icc_ccc(dp)
                if len(dp) < 5:
                    continue
                icc_vals = []
                ccc_vals = []
                for pid in sorted(dp.pid.unique()):
                    dp_loo = dp[dp.pid != pid].copy()
                    loo_val = icc_ccc(dp_loo)
                    d_icc = full["ICC3k"] - loo_val["ICC3k"]
                    d_ccc = full["CCC"] - loo_val["CCC"]
                    icc_vals.append(loo_val["ICC3k"])
                    ccc_vals.append(loo_val["CCC"])
                    loo_rows.append({
                        "Group": group,
                        "Variable": label,
                        "Epoch": EPOCH_LABEL[epoch],
                        "pid_left_out": pid,
                        "n_full": len(dp),
                        "ICC_full": full["ICC3k"],
                        "CCC_full": full["CCC"],
                        "ICC_leave_one_out": loo_val["ICC3k"],
                        "CCC_leave_one_out": loo_val["CCC"],
                        "delta_ICC_full_minus_LOO": d_icc,
                        "delta_CCC_full_minus_LOO": d_ccc,
                        "abs_delta_ICC": abs(d_icc),
                        "abs_delta_CCC": abs(d_ccc),
                    })
                summary_rows.append({
                    "Group": group,
                    "Variable": label,
                    "Epoch": EPOCH_LABEL[epoch],
                    "n": len(dp),
                    "ICC_full": full["ICC3k"],
                    "CCC_full": full["CCC"],
                    "ICC_LOO_min": np.nanmin(icc_vals),
                    "ICC_LOO_max": np.nanmax(icc_vals),
                    "CCC_LOO_min": np.nanmin(ccc_vals),
                    "CCC_LOO_max": np.nanmax(ccc_vals),
                    "max_abs_delta_ICC": np.nanmax(np.abs(np.array(icc_vals) - full["ICC3k"])),
                    "max_abs_delta_CCC": np.nanmax(np.abs(np.array(ccc_vals) - full["CCC"])),
                })

    summary = pd.DataFrame(summary_rows)
    loo = pd.DataFrame(loo_rows)

    idx_icc = loo.groupby(["Group", "Variable", "Epoch"])["abs_delta_ICC"].idxmax()
    max_icc = loo.loc[idx_icc, [
        "Group", "Variable", "Epoch", "pid_left_out", "delta_ICC_full_minus_LOO",
        "abs_delta_ICC", "ICC_leave_one_out"
    ]].rename(columns={
        "pid_left_out": "Most_influential_ID_ICC",
        "delta_ICC_full_minus_LOO": "ICC_change_full_minus_leave_one_out",
        "abs_delta_ICC": "Max_abs_change_ICC",
        "ICC_leave_one_out": "ICC_after_removing_most_influential_ID",
    })
    idx_ccc = loo.groupby(["Group", "Variable", "Epoch"])["abs_delta_CCC"].idxmax()
    max_ccc = loo.loc[idx_ccc, [
        "Group", "Variable", "Epoch", "pid_left_out", "delta_CCC_full_minus_LOO",
        "abs_delta_CCC", "CCC_leave_one_out"
    ]].rename(columns={
        "pid_left_out": "Most_influential_ID_CCC",
        "delta_CCC_full_minus_LOO": "CCC_change_full_minus_leave_one_out",
        "abs_delta_CCC": "Max_abs_change_CCC",
        "CCC_leave_one_out": "CCC_after_removing_most_influential_ID",
    })
    out = summary.merge(max_icc, on=["Group", "Variable", "Epoch"], how="left")
    out = out.merge(max_ccc, on=["Group", "Variable", "Epoch"], how="left")
    out["ICC_sign_changes_in_LOO"] = (out["ICC_LOO_min"] < 0) & (out["ICC_LOO_max"] > 0)
    out["CCC_sign_changes_in_LOO"] = (out["CCC_LOO_min"] < 0) & (out["CCC_LOO_max"] > 0)
    out["High_influence_flag"] = (out["Max_abs_change_ICC"] >= 0.30) | (out["Max_abs_change_CCC"] >= 0.20)
    for col in out.select_dtypes(include="number").columns:
        if col not in ["n", "Most_influential_ID_ICC", "Most_influential_ID_CCC"]:
            out[col] = out[col].round(3)
    out = out.sort_values(["Group", "Variable", "Epoch"])
    out.to_csv(FIG_DIR / "SupplementaryTable_Reliability_Influence_LOO.csv", index=False)
    flagged = out[out["High_influence_flag"]].copy()
    flagged.to_csv(FIG_DIR / "SupplementaryTable_Reliability_Influence_LOO_Flagged.csv", index=False)
    print(f"Saved: SupplementaryTable_Reliability_Influence_LOO.csv ({len(out)} rows)")
    print(f"Saved: SupplementaryTable_Reliability_Influence_LOO_Flagged.csv ({len(flagged)} rows)")


def regression_table(df):
    rows = []
    for _, row in df.iterrows():
        pid = int(row["1 Identifier"])
        sid = str(row.get("1 Subject ID", ""))
        if pid in GLOBAL_EXCL_IDS:
            continue
        for visit in (1, 2):
            if (pid, visit) in VISIT_EXCL:
                continue
            rows.append({
                "pid": pid,
                "sid": sid,
                "visit": visit,
                "delta_cvci": to_num(row.get(f"{visit} Delta MCAv CVCi CPT 2min")),
                "pain": to_num(row.get(f"{visit} rating of discomfort")),
                "delta_etco2": to_num(row.get(f"{visit} Delta ET-CO2 CPT 2min")),
            })
    lmm = pd.DataFrame(rows)
    cc = lmm.dropna(subset=["delta_cvci", "pain", "delta_etco2"]).copy()
    model = smf.mixedlm("delta_cvci ~ pain + delta_etco2 + visit", cc, groups=cc["pid"])
    last_exc = None
    for method in ("lbfgs", "bfgs", "nm"):
        try:
            res = model.fit(reml=True, method=method, disp=False)
            optimizer = method
            break
        except Exception as exc:
            last_exc = exc
    else:
        raise last_exc
    sd_y = cc["delta_cvci"].std(ddof=1)
    beta = {
        "Intercept": np.nan,
        "Perceived Pain (VAS)": res.params["pain"] * cc["pain"].std(ddof=1) / sd_y,
        "Delta ET-CO2 (mmHg)": res.params["delta_etco2"] * cc["delta_etco2"].std(ddof=1) / sd_y,
        "Visit Number": np.nan,
    }
    name_map = {
        "Intercept": "Intercept",
        "pain": "Perceived Pain (VAS)",
        "delta_etco2": "Delta ET-CO2 (mmHg)",
        "visit": "Visit Number",
    }
    out_rows = []
    for param, label in name_map.items():
        out_rows.append({
            "Predictor": label,
            "b": res.params[param],
            "beta": beta[label],
            "SE": res.bse[param],
            "t": res.tvalues[param],
            "P": res.pvalues[param],
            "n_participants": cc["pid"].nunique(),
            "n_observations": len(cc),
            "optimizer": optimizer,
        })
    out = pd.DataFrame(out_rows)
    out.to_csv(FIG_DIR / "Table4_Regression_Updated_ETCO2.csv", index=False)
    cc.to_csv(FIG_DIR / "Table4_Regression_Analytic_Data.csv", index=False)
    print("Saved: Table4_Regression_Updated_ETCO2.csv")
    print("Saved: Table4_Regression_Analytic_Data.csv")
    print(f"Regression analytic sample: {cc['pid'].nunique()} participants, {len(cc)} observations.")
    print(out[["Predictor", "b", "beta", "SE", "t", "P"]].to_string(index=False))


def main():
    df = load_data()
    fig1_timecourse(df)
    fig1_anova(df)
    stats_df = build_tables_and_stats(df)
    ba_grid(df, FIG2_VARS, "Fig2_BA_Cardiovascular.png", "Bland-Altman plots: cardiovascular variables")
    ba_grid(df, FIG3_VARS, "Fig3_BA_Cerebrovascular.png", "Bland-Altman plots: cerebrovascular variables")
    reliability_summary_plot(stats_df, FIG2_VARS + FIG3_VARS, "Fig4_ICC_CCC_Summary.png", "ICC and CCC by epoch")
    sex_df = sex_reliability(df, FIG2_VARS + FIG3_VARS)
    sex_reliability_plot(sex_df, FIG2_VARS + FIG3_VARS)
    fig6_stats = build_reliability_stats_for_vars(df, FIG2_VARS + FIG3_VARS, subset=exclude_unmatched_females)
    fig6_stats.to_csv(FIG_DIR / "Fig6_ICC_CCC_Exclude_Unmatched_Females_Stats.csv", index=False)
    print("Saved: Fig6_ICC_CCC_Exclude_Unmatched_Females_Stats.csv")
    reliability_summary_plot(
        fig6_stats,
        FIG2_VARS + FIG3_VARS,
        "Fig6_ICC_CCC_Exclude_Unmatched_Females.png",
        "ICC and CCC by epoch: unmatched females excluded",
    )
    export_loo_influence_table(df, FIG2_VARS + FIG3_VARS)
    save_png_as_pdf(FIG_DIR / "SupFig_ICC_CCC_AllVars.png")
    print("Done.")


if __name__ == "__main__":
    main()
